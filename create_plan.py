# -*- coding: utf-8 -*-
"""5年生 年間単元指導計画表 Excel生成スクリプト"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "5年年間単元指導計画"

# ── ページ設定（A3横向き） ──
ws.page_setup.paperSize = 8  # A3
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.4
ws.page_margins.right = 0.4
ws.page_margins.top = 0.4
ws.page_margins.bottom = 0.4

# ── 定数 ──
SUBJECTS = [
    "国語", "算数", "社会", "理科", "音楽", "図画工作",
    "体育", "家庭", "外国語", "道徳・特活", "健康", "食育",
    "自己実現活動", "安全",
]
MONTHS = ["4月","5月","6月","7月","8月","9月","10月","11月","12月","1月","2月","3月"]

SEMESTER = {  # 学期判定: 月 → 学期番号
    "4月":1,"5月":1,"6月":1,"7月":1,"8月":1,
    "9月":2,"10月":2,"11月":2,"12月":2,
    "1月":3,"2月":3,"3月":3,
}

# ── スタイル ──
FONT      = Font(name="游ゴシック", size=9)
FONT_B    = Font(name="游ゴシック", size=9, bold=True)
FONT_TTL  = Font(name="游ゴシック", size=16, bold=True)
FONT_TOT  = Font(name="游ゴシック", size=10, bold=True)
FONT_ANN  = Font(name="游ゴシック", size=9, bold=True, color="333333")

ALIGN_C   = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_TL  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
ALIGN_CR  = Alignment(horizontal="center", vertical="center")

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_HDR = PatternFill("solid", fgColor="D5C8E8")   # ヘッダー紫
FILL_1   = PatternFill("solid", fgColor="DCE6F1")    # 1学期 薄い青
FILL_2   = PatternFill("solid", fgColor="FFF2CC")    # 2学期 薄い黄
FILL_3   = PatternFill("solid", fgColor="D5F5E3")    # 3学期 薄い緑
FILL_TOT = PatternFill("solid", fgColor="E8E0F0")    # 合計行
FILL_SEM = {1: FILL_1, 2: FILL_2, 3: FILL_3}

# ── 単元テキストデータ ──
TEXT = {s: {m: "" for m in MONTHS} for s in SUBJECTS}
HOURS = {s: {m: 0 for m in MONTHS} for s in SUBJECTS}

def add(subj, month, name, h):
    if TEXT[subj][month]:
        TEXT[subj][month] += "\n"
    TEXT[subj][month] += f"{name}({h})"
    HOURS[subj][month] += h

# ── 国語 174h ──
add("国語","4月","ひみつの言葉を引き出そう",1)
add("国語","4月","かんがえるのって おもしろい",1)
add("国語","4月","銀色の裏地",5)
add("国語","4月","図書館をつかいこなそう",1)
add("国語","4月","春の空",2)
add("国語","4月","漢字の広場①・漢字の成り立ち・敬語・同じ読み方の漢字",7)
add("国語","5月","名前を使って自己しょうかい",1)
add("国語","5月","きいて、きいて、きいてみよう",6)
add("国語","5月","見立てる／言葉の意味が分かること・原因と結果",7)
add("国語","5月","毛筆・硬筆",30)
add("国語","6月","日常を十七音で",3)
add("国語","6月","古典の世界(一)",2)
add("国語","6月","目的に応じて引用する時",2)
add("国語","6月","みんなが使いやすいデザイン",8)
add("国語","7月","夏の夜",1)
add("国語","7月","作家で広げるわたしたちの読書【モモ】",5)
add("国語","9月","かぼちゃのつるが・われは草なり",2)
add("国語","9月","どちらを選びますか",2)
add("国語","9月","新聞を読もう",2)
add("国語","9月","文章に説得力をもたせるには",2)
add("国語","9月","たずねびと",6)
add("国語","10月","秋の夕",1)
add("国語","10月","よりよい学校生活のために",6)
add("国語","11月","浦島太郎―「御伽草子」より",1)
add("国語","11月","固有種が教えてくれること・統計資料の読み方",10)
add("国語","11月","古典の世界(二)",1)
add("国語","12月","やなせたかし―アンパンマンの勇気",5)
add("国語","12月","あなたは、どう考える",6)
add("国語","12月","冬の朝",1)
add("国語","1月","好きな詩のよさを伝えよう",2)
add("国語","1月","言葉でスケッチ",2)
add("国語","1月","想像力のスイッチを入れよう",6)
add("国語","1月","熟語の読み方・漢字の広場⑤⑥・複合語",6)
add("国語","2月","方言と共通語・漢字の広場②〜④・和語漢語外来語他",9)
add("国語","2月","伝わる表現を選ぼう",3)
add("国語","2月","もう一つの物語",6)
add("国語","2月","子ども未来科で何する",6)
add("国語","3月","大造じいさんとがん",6)
add("国語","3月","五年生をふり返って",1)

# ── 算数 156h ──
add("算数","4月","整数と小数",5)
add("算数","4月","直方体や立方体の体積",11)
add("算数","5月","比例",3)
add("算数","5月","小数のかけ算",12)
add("算数","6月","小数のわり算",14)
add("算数","7月","合同な図形",9)
add("算数","8月","偶数と奇数、倍数と約数",11)
add("算数","9月","分数と小数、整数の関係",6)
add("算数","10月","考える力をのばそう",2)
add("算数","10月","分数のたし算とひき算",12)
add("算数","11月","単位量あたりの大きさ",13)
add("算数","11月","速さ",7)
add("算数","12月","四角形と三角形の面積",13)
add("算数","12月","算数の目で見てみよう",1)
add("算数","1月","図形の角",7)
add("算数","1月","百分率とグラフ",12)
add("算数","2月","考える力をのばそう(表を利用して)",1)
add("算数","2月","正多角形と円周の長さ",10)
add("算数","3月","角柱と円柱",5)
add("算数","3月","算数の目で見てみよう(資料の活用)",2)

# ── 社会 115h ──
add("社会","4月","国土の様子(大陸と海洋・位置と領土)",40)
add("社会","4月","我が国の食料生産",20)
add("社会","9月","我が国の工業生産",20)
add("社会","11月","我が国の情報と産業との関わり",20)
add("社会","1月","国土の自然環境と国民生活との関連",15)

# ── 理科 86h ──
add("理科","4月","天気の変化",6)
add("理科","4月","植物の発芽成長",15)
add("理科","6月","メダカのたんじょう",6)
add("理科","7月","植物の実や種子のでき方",4)
add("理科","9月","流れる水のはたらきと土地の変化",6)
add("理科","9月","ふりこ",9)
add("理科","10月","電流が生み出す力",10)
add("理科","1月","人のたんじょう",6)
add("理科","2月","ものの溶け方",15)
add("理科","2月","電磁石の性質",9)

# ── 音楽 24h ──
add("音楽","4月","オーケストラの秘密",6)
add("音楽","4月","みんなでつくろう",6)
add("音楽","9月","みんなでつくろう",6)
add("音楽","1月","変化を生かして合奏をつくろう",6)

# ── 図画工作 48h ──
add("図画工作","4月","アート造形",10)
add("図画工作","6月","アートお絵描き(自分を見つめて)",4)
add("図画工作","6月","1枚の板から",8)
add("図画工作","9月","竹早祭展覧会(アート造形)",8)
add("図画工作","10月","ガラスアート(陶芸)",4)
add("図画工作","11月","浮世絵の世界(木版画)",6)
add("図画工作","1月","アート造形",8)

# ── 体育 76h ──
add("体育","4月","短距離走",4)
add("体育","4月","体の動きを高める運動",4)
add("体育","5月","表現(竹の子祭)",4)
add("体育","5月","竹の子祭個人種目",4)
add("体育","6月","体ほぐしの運動",4)
add("体育","6月","リレー",4)
add("体育","6月","走り高跳び",4)
add("体育","7月","クロール＋平泳ぎ＋安全確保につながる運動",8)
add("体育","9月","マット運動",4)
add("体育","9月","ゴール型ゲーム",4)
add("体育","10月","新体力テスト",4)
add("体育","10月","ハードル走",4)
add("体育","10月","ネット型ゲーム",4)
add("体育","11月","跳び箱運動",4)
add("体育","11月","ゴール型ゲーム",4)
add("体育","1月","体の動きを高める運動",4)
add("体育","1月","ゴール型ゲーム(体育館)",4)
add("体育","1月","ゴール型ゲーム(校庭)",4)

# ── 家庭 36h ──
add("家庭","4月","一日の生活をふりかえってみよう",4)
add("家庭","5月","かんたんな仕事をしてみよう",4)
add("家庭","6月","かんたんな調理をしてみよう",4)
add("家庭","7月","環境を考えた後片付けをしよう",4)
add("家庭","9月","針と糸を使ってみよう",4)
add("家庭","10月","ミシンを使ってみよう",4)
add("家庭","11月","小物を作ろう",4)
add("家庭","12月","アイロンを使おう→作品展示",4)
add("家庭","1月","快適な住まい方の工夫をしよう",4)

# ── 外国語 54h ──
add("外国語","4月","自己紹介",6)
add("外国語","5月","数・値段",6)
add("外国語","6月","教科・習い事",6)
add("外国語","9月","位置・道案内",5)
add("外国語","10月","一日の生活",5)
add("外国語","11月","できること",5)
add("外国語","12月","誕生日・ほしいもの",5)
add("外国語","1月","食べ物の注文",5)
add("外国語","2月","国・地域",5)
add("外国語","3月","あこがれの人",6)

# ── 道徳・特活 20h ──
add("道徳・特活","4月","B 他者との関わりの追究\n(親切・思いやり・感謝・礼儀・友情信頼・相互理解寛容)\nD 環境の探究(通年)",7)
add("道徳・特活","9月","C 集団との関わりの追究\n(規則の尊重・公正公平・勤労・家族愛・学校生活・伝統文化・国際理解)\nD 環境の探究(通年)",7)
add("道徳・特活","1月","A 自己の確立\n(善悪の判断・正直誠実・節度節制・個性の伸長・希望と勇気・真理の探究)\nD 環境の探究(通年)",6)

# ── 健康 28h ──
add("健康","4月","もれなく受けよう健康診断",4)
add("健康","10月","発育測定",4)
add("健康","11月","ケガしたマップをつくろう",4)
add("健康","12月","健康リテラシーについて考えよう",4)
add("健康","2月","心の健康",4)
add("健康","3月","発育測定",4)
add("健康","3月","ケガしたマップをつくろう",4)

# ── 食育 8h ──
add("食育","5月","バランスのよい食事をしよう",4)
add("食育","11月","留学生交流会(日本の食文化について)",4)

# ── 自己実現活動 16h ──
add("自己実現活動","6月","竹の子祭",4)
add("自己実現活動","10月","秋の日光",4)
add("自己実現活動","11月","竹早祭",4)
add("自己実現活動","3月","６送会",4)

# ── 安全 24h ──
add("安全","6月","学校生活をふりかえって",4)
add("安全","7月","セーフティ教室",4)
add("安全","12月","学校生活をふりかえって",4)
add("安全","1月","不審者訓練",4)
add("安全","2月","小中高合同避難訓練",4)
add("安全","3月","学校生活をふりかえって",4)

# ══════════════════════════════════════
#  Excelシート構築
# ══════════════════════════════════════
NUM_COLS = 1 + len(SUBJECTS)  # A=月, B〜O=教科

# ---- Row 1: タイトル ----
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
title_cell = ws.cell(row=1, column=1, value="令和7年度　第5学年　年間単元指導計画表")
title_cell.font = FONT_TTL
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# ---- Row 2: ヘッダー ----
headers = ["月"] + SUBJECTS
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=2, column=ci, value=h)
    c.font = FONT_B
    c.alignment = ALIGN_C
    c.border = BORDER
    c.fill = FILL_HDR
ws.row_dimensions[2].height = 32

# ---- Rows 3〜26: 月データ (text行 + 非表示hours行 × 12か月) ----
for mi, month in enumerate(MONTHS):
    text_row = 3 + mi * 2
    hour_row = 4 + mi * 2
    sem = SEMESTER[month]
    fill = FILL_SEM[sem]

    # --- テキスト行 ---
    # 月セル
    mc = ws.cell(row=text_row, column=1, value=month)
    mc.font = FONT_B
    mc.alignment = ALIGN_C
    mc.border = BORDER
    mc.fill = fill

    for si, subj in enumerate(SUBJECTS):
        col = si + 2
        txt = TEXT[subj][month]
        c = ws.cell(row=text_row, column=col, value=txt if txt else "")
        c.font = FONT
        c.alignment = ALIGN_TL
        c.border = BORDER
        c.fill = fill

    # テキスト行の高さ（内容量に応じて調整）
    max_lines = 1
    for subj in SUBJECTS:
        txt = TEXT[subj][month]
        if txt:
            max_lines = max(max_lines, txt.count("\n") + 1)
    ws.row_dimensions[text_row].height = max(36, min(max_lines * 14, 120))

    # --- 非表示 hours 行（SUM計算用） ---
    hmc = ws.cell(row=hour_row, column=1, value="")
    hmc.border = BORDER
    for si, subj in enumerate(SUBJECTS):
        col = si + 2
        h = HOURS[subj][month]
        c = ws.cell(row=hour_row, column=col, value=h if h > 0 else 0)
        c.font = FONT
        c.alignment = ALIGN_CR
        c.border = BORDER
        c.number_format = "0"
    ws.row_dimensions[hour_row].hidden = True  # 非表示

# ---- Row 27: 年間合計 ----
TOTAL_ROW = 3 + 12 * 2  # = 27
tc = ws.cell(row=TOTAL_ROW, column=1, value="年間合計")
tc.font = FONT_TOT
tc.alignment = ALIGN_C
tc.border = BORDER
tc.fill = FILL_TOT

for si, subj in enumerate(SUBJECTS):
    col = si + 2
    col_letter = get_column_letter(col)
    # SUM: 非表示行(4,6,8,...,26)の合計
    hour_cells = ",".join(f"{col_letter}{4 + i*2}" for i in range(12))
    formula = f"=SUM({hour_cells})"
    c = ws.cell(row=TOTAL_ROW, column=col)
    c.value = formula
    c.font = FONT_TOT
    c.alignment = ALIGN_CR
    c.border = BORDER
    c.fill = FILL_TOT
    c.number_format = "0"

ws.row_dimensions[TOTAL_ROW].height = 28

# ---- 年間時数 ラベル行 ----
ANNO_ROW = TOTAL_ROW + 1
ws.merge_cells(start_row=ANNO_ROW, start_column=1, end_row=ANNO_ROW, end_column=NUM_COLS)
ac = ws.cell(row=ANNO_ROW, column=1,
    value="※ 社会・音楽等の大単元は開始月に年間時数をまとめて記載しています。実際の配当は月ごとに分散されます。")
ac.font = Font(name="游ゴシック", size=8, color="666666")
ac.alignment = Alignment(horizontal="left", vertical="center")

# ── 列幅設定（A3横向き印刷用） ──
ws.column_dimensions["A"].width = 5.5   # 月
WIDE_SUBJECTS = {"国語": 22, "算数": 18, "社会": 18, "理科": 16, "道徳・特活": 18}
DEFAULT_WIDTH = 14

for si, subj in enumerate(SUBJECTS):
    col_letter = get_column_letter(si + 2)
    ws.column_dimensions[col_letter].width = WIDE_SUBJECTS.get(subj, DEFAULT_WIDTH)

# ── 印刷設定 ──
ws.print_title_rows = "1:2"  # 各ページにヘッダーを印刷

# ── 保存 ──
OUTPUT = r"C:\Users\nkmr2\gakkyu-keiei\5年_年間単元指導計画表.xlsx"
wb.save(OUTPUT)
print(f"保存完了: {OUTPUT}")

# 検証: 年間合計の確認
print("\n【年間合計時数の検証】")
for subj in SUBJECTS:
    total = sum(HOURS[subj][m] for m in MONTHS)
    print(f"  {subj}: {total}時間")
