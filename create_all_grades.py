# -*- coding: utf-8 -*-
"""全学年の年間単元指導計画 確認用Excelファイルを一括生成するスクリプト

使い方:
    python create_all_grades.py

出力:
    excel_templates/ ディレクトリに学年ごとのExcelファイルを生成
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ── データ読み込み ──
DATA_PATH = Path(__file__).parent / "curriculum_data.json"
with open(DATA_PATH, "r", encoding="utf-8") as f:
    CURRICULUM = json.load(f)

OUTPUT_DIR = Path(__file__).parent / "excel_templates"
OUTPUT_DIR.mkdir(exist_ok=True)

# ── 定数 ──
ALL_SUBJECTS = [
    "国語", "算数", "社会", "理科", "音楽", "図画工作",
    "体育", "家庭", "外国語", "道徳・特活", "情報教育",
    "健康", "食育", "自己実現活動", "安全",
]
MONTHS = ["4月","5月","6月","7月","8月","9月","10月","11月","12月","1月","2月","3月"]
SEMESTER = {
    "4月":1,"5月":1,"6月":1,"7月":1,"8月":1,
    "9月":2,"10月":2,"11月":2,"12月":2,
    "1月":3,"2月":3,"3月":3,
}
MONTH_NUM = {"4月":4,"5月":5,"6月":6,"7月":7,"8月":8,"9月":9,
             "10月":10,"11月":11,"12月":12,"1月":1,"2月":2,"3月":3}

# ── スタイル ──
FONT      = Font(name="游ゴシック", size=9)
FONT_B    = Font(name="游ゴシック", size=9, bold=True)
FONT_TTL  = Font(name="游ゴシック", size=14, bold=True)
FONT_TOT  = Font(name="游ゴシック", size=10, bold=True)
FONT_NOTE = Font(name="游ゴシック", size=8, color="666666")

ALIGN_C   = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_TL  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
ALIGN_CR  = Alignment(horizontal="center", vertical="center")

THIN   = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_HDR = PatternFill("solid", fgColor="D5C8E8")
FILL_1   = PatternFill("solid", fgColor="DCE6F1")
FILL_2   = PatternFill("solid", fgColor="FFF2CC")
FILL_3   = PatternFill("solid", fgColor="D5F5E3")
FILL_TOT = PatternFill("solid", fgColor="E8E0F0")
FILL_SEM = {1: FILL_1, 2: FILL_2, 3: FILL_3}

WIDE_SUBJECTS = {"国語": 22, "算数": 18, "社会": 18, "理科": 16, "道徳・特活": 18}
DEFAULT_WIDTH = 14


def create_grade_excel(grade_str):
    """1学年分のExcelを生成"""
    grade_data = CURRICULUM[grade_str]
    # この学年で使う教科だけフィルタ
    subjects = [s for s in ALL_SUBJECTS if s in grade_data]

    # 単元テキストと時数を構築
    TEXT  = {s: {m: "" for m in MONTHS} for s in subjects}
    HOURS = {s: {m: 0  for m in MONTHS} for s in subjects}

    for subj in subjects:
        for unit in grade_data[subj]:
            m_num = unit["month"]
            m_label = f"{m_num}月"
            if m_label not in TEXT[subj]:
                continue
            name = unit["unit"]
            h = unit["hours"]
            if TEXT[subj][m_label]:
                TEXT[subj][m_label] += "\n"
            TEXT[subj][m_label] += f"{name}({h})"
            HOURS[subj][m_label] += h

    # ワークブック作成
    wb = Workbook()
    ws = wb.active
    ws.title = f"{grade_str}年年間単元指導計画"

    # ページ設定（A3横向き）
    ws.page_setup.paperSize = 8
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    NUM_COLS = 1 + len(subjects)

    # ── Row 1: タイトル ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    title_cell = ws.cell(row=1, column=1,
        value=f"令和7年度　第{grade_str}学年　年間単元指導計画表【確認用】")
    title_cell.font = FONT_TTL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # ── Row 2: ヘッダー ──
    headers = ["月"] + subjects
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = FONT_B
        c.alignment = ALIGN_C
        c.border = BORDER
        c.fill = FILL_HDR
    ws.row_dimensions[2].height = 32

    # ── Rows 3〜26: 月データ ──
    for mi, month in enumerate(MONTHS):
        text_row = 3 + mi * 2
        hour_row = 4 + mi * 2
        sem = SEMESTER[month]
        fill = FILL_SEM[sem]

        # テキスト行
        mc = ws.cell(row=text_row, column=1, value=month)
        mc.font = FONT_B
        mc.alignment = ALIGN_C
        mc.border = BORDER
        mc.fill = fill

        for si, subj in enumerate(subjects):
            col = si + 2
            txt = TEXT[subj][month]
            c = ws.cell(row=text_row, column=col, value=txt if txt else "")
            c.font = FONT
            c.alignment = ALIGN_TL
            c.border = BORDER
            c.fill = fill

        # テキスト行の高さ
        max_lines = 1
        for subj in subjects:
            txt = TEXT[subj][month]
            if txt:
                max_lines = max(max_lines, txt.count("\n") + 1)
        ws.row_dimensions[text_row].height = max(36, min(max_lines * 14, 120))

        # 非表示 hours 行
        hmc = ws.cell(row=hour_row, column=1, value="")
        hmc.border = BORDER
        for si, subj in enumerate(subjects):
            col = si + 2
            h = HOURS[subj][month]
            c = ws.cell(row=hour_row, column=col, value=h if h > 0 else 0)
            c.font = FONT
            c.alignment = ALIGN_CR
            c.border = BORDER
            c.number_format = "0"
        ws.row_dimensions[hour_row].hidden = True

    # ── Row 27: 年間合計 ──
    TOTAL_ROW = 3 + 12 * 2
    tc = ws.cell(row=TOTAL_ROW, column=1, value="年間合計")
    tc.font = FONT_TOT
    tc.alignment = ALIGN_C
    tc.border = BORDER
    tc.fill = FILL_TOT

    for si, subj in enumerate(subjects):
        col = si + 2
        col_letter = get_column_letter(col)
        hour_cells = ",".join(f"{col_letter}{4 + i*2}" for i in range(12))
        formula = f"=SUM({hour_cells})"
        c = ws.cell(row=TOTAL_ROW, column=col)
        c.value = formula
        c.font = FONT_TOT
        c.alignment = ALIGN_CR
        c.border = BORDER
        c.fill = FILL_TOT
        c.number_format = "0"

    ws.row_dimensions[TOTAL_ROW].height = 28

    # ── 注釈行 ──
    ANNO_ROW = TOTAL_ROW + 1
    ws.merge_cells(start_row=ANNO_ROW, start_column=1, end_row=ANNO_ROW, end_column=NUM_COLS)
    ac = ws.cell(row=ANNO_ROW, column=1,
        value="※ 先生方へ：各教科の単元名と時数をご確認ください。修正がある場合はセルを直接編集してください。書式は「単元名(時数)」です。")
    ac.font = FONT_NOTE
    ac.alignment = Alignment(horizontal="left", vertical="center")

    # ── 列幅設定 ──
    ws.column_dimensions["A"].width = 5.5
    for si, subj in enumerate(subjects):
        col_letter = get_column_letter(si + 2)
        ws.column_dimensions[col_letter].width = WIDE_SUBJECTS.get(subj, DEFAULT_WIDTH)

    # ── 印刷設定 ──
    ws.print_title_rows = "1:2"

    # ── 保存 ──
    output_path = OUTPUT_DIR / f"{grade_str}年_年間単元指導計画表_確認用.xlsx"
    wb.save(str(output_path))
    return output_path


def main():
    print("年間単元指導計画 確認用Excelファイルを生成中...\n")
    for grade in sorted(CURRICULUM.keys(), key=int):
        path = create_grade_excel(grade)
        # 年間合計の検証
        grade_data = CURRICULUM[grade]
        subjects = [s for s in ALL_SUBJECTS if s in grade_data]
        print(f"  {grade}年生: {path.name}  ({len(subjects)}教科)")
        for subj in subjects:
            total = sum(u["hours"] for u in grade_data[subj])
            if total > 0:
                print(f"    {subj}: {total}時間")
            else:
                print(f"    {subj}: （未入力）")
    print(f"\n完了: {OUTPUT_DIR}/")


if __name__ == "__main__":
    main()
